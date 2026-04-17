#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Generate frontend product data from 产品识别.xlsx."""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import List, Dict, Tuple

from openpyxl import load_workbook

PROJECT_ROOT = Path(r"C:\Users\Win11\Desktop\XCSOURCE")
PRODUCTS_DIR = PROJECT_ROOT / "assets" / "products"
EXCEL_PATH = PRODUCTS_DIR / "产品识别.xlsx"
OUTPUT_JSON = PRODUCTS_DIR / "products.json"
REQUIRED_HEADERS = ["是否上传", "图片超链接", "英文描述"]
HYPERLINK_RE = re.compile(r'=HYPERLINK\("([^"]+)"\s*,\s*"([^"]*)"\)', re.IGNORECASE)


def normalize_header(value: object) -> str:
    return str(value).strip() if value is not None else ""


def parse_hyperlink(cell) -> Tuple[str | None, str | None]:
    if cell.hyperlink and cell.hyperlink.target:
        target = str(cell.hyperlink.target).strip()
        display = str(cell.value).strip() if cell.value is not None else ""
        return target, display

    raw = str(cell.value).strip() if cell.value is not None else ""
    match = HYPERLINK_RE.fullmatch(raw)
    if match:
        return match.group(1).strip(), match.group(2).strip()

    return None, None


def to_web_path(path: Path) -> str:
    return path.relative_to(PROJECT_ROOT).as_posix()


def main() -> int:
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"产品识别.xlsx 不存在: {EXCEL_PATH}")

    wb = load_workbook(EXCEL_PATH, data_only=False)
    ws = wb[wb.sheetnames[0]]

    headers = [normalize_header(cell.value) for cell in ws[1]]
    header_map = {header: index for index, header in enumerate(headers)}
    missing = [header for header in REQUIRED_HEADERS if header not in header_map]
    if missing:
        raise KeyError(f"缺少必要字段: {', '.join(missing)}")

    stats: Dict[str, object] = {
        "records_read": 0,
        "valid_products": 0,
        "written_products": 0,
        "skipped_rows": 0,
        "exceptions": [],
    }
    products: List[Dict[str, str]] = []

    for row_number in range(2, ws.max_row + 1):
        stats["records_read"] += 1

        upload_value = ws.cell(row_number, header_map["是否上传"] + 1).value
        upload_value = str(upload_value).strip() if upload_value is not None else ""

        image_cell = ws.cell(row_number, header_map["图片超链接"] + 1)
        description_value = ws.cell(row_number, header_map["英文描述"] + 1).value
        description = str(description_value).strip() if description_value is not None else ""

        if not upload_value:
            stats["skipped_rows"] += 1
            stats["exceptions"].append(f"第{row_number}行: 是否上传为空")
            continue

        if upload_value != "否":
            continue

        link_target, display_text = parse_hyperlink(image_cell)
        if not link_target:
            stats["skipped_rows"] += 1
            stats["exceptions"].append(f"第{row_number}行: 图片超链接无效")
            continue

        if not description:
            stats["skipped_rows"] += 1
            stats["exceptions"].append(f"第{row_number}行: 英文描述为空")
            continue

        link_path = Path(link_target.replace('\\', '/'))
        if link_path.is_absolute():
            image_path = link_path
        else:
            image_path = PRODUCTS_DIR / link_path

        if not image_path.exists() or not image_path.is_file():
            stats["skipped_rows"] += 1
            stats["exceptions"].append(f"第{row_number}行: 图片文件不存在 -> {link_target}")
            continue

        if PROJECT_ROOT not in image_path.parents and image_path != PROJECT_ROOT:
            stats["skipped_rows"] += 1
            stats["exceptions"].append(f"第{row_number}行: 图片路径超出项目目录 -> {image_path}")
            continue

        products.append(
            {
                "image": to_web_path(image_path),
                "description": description,
                "sourceRow": row_number,
                "imageName": display_text or image_path.name,
            }
        )
        stats["valid_products"] += 1

    OUTPUT_JSON.write_text(json.dumps(products, ensure_ascii=False, indent=2), encoding="utf-8")
    stats["written_products"] = len(products)

    print(json.dumps(stats, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
