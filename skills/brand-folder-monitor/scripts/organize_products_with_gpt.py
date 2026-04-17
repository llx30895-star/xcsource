#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
产品图片识别和表格整理脚本（使用 OpenAI GPT-5.4 Vision + 自动转移）
使用 OpenAI Vision API 识别产品图片，生成提示词和中英文描述
识别完成后自动将图片转移到【已上传图片】文件夹
"""

import os
import sys
import shutil
import subprocess
from pathlib import Path
from datetime import datetime
import json
import base64
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

try:
    from openai import OpenAI
except ImportError:
    print("错误: 需要安装 openai 库")
    print("运行: pip install openai")
    sys.exit(1)

# 配置
BASE_DIR = Path(r"C:\Users\Win11\Desktop\XCSOURCE\assets\products")
UPLOADED_DIR = BASE_DIR / "已上传图片"
EXCEL_FILE = BASE_DIR / "产品识别.xlsx"
GENERATE_SCRIPT = BASE_DIR.parent.parent / "generate_products_data.py"
IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp"}
LOG_FILE = BASE_DIR / "产品识别.log"

# API 配置
API_KEY = os.environ.get("OPENAI_API_KEY", "")
MODEL = "gpt-5.4"


def log(msg):
    """写日志"""
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(line + "\n")


def init_excel():
    """初始化 Excel 表格"""
    if EXCEL_FILE.exists():
        log(f"Excel 文件已存在: {EXCEL_FILE}")
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ensure_headers(ws)
        wb.save(EXCEL_FILE)
        return wb

    log(f"创建新的 Excel 文件: {EXCEL_FILE}")
    wb = Workbook()
    ws = wb.active
    ws.title = "产品识别"

    headers = [
        "序号", "添加时间", "图片超链接", "提示词", "中文描述", "英文描述", "是否上传",
        "记录ID", "流程阶段", "识别状态", "前端生成状态", "上传状态", "上传时间",
        "上传目标", "上传批次号", "上传结果", "失败原因", "最近处理时间"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 12
    ws.column_dimensions['M'].width = 18
    ws.column_dimensions['N'].width = 15
    ws.column_dimensions['O'].width = 20
    ws.column_dimensions['P'].width = 25
    ws.column_dimensions['Q'].width = 30
    ws.column_dimensions['R'].width = 18

    wb.save(EXCEL_FILE)
    return wb


def ensure_headers(ws):
    """确保工作表包含完整表头，如果缺失字段则自动追加"""
    expected_headers = [
        "序号", "添加时间", "图片超链接", "提示词", "中文描述", "英文描述", "是否上传",
        "记录ID", "流程阶段", "识别状态", "前端生成状态", "上传状态", "上传时间",
        "上传目标", "上传批次号", "上传结果", "失败原因", "最近处理时间"
    ]
    
    existing_headers = []
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(1, col).value
        if cell_value:
            existing_headers.append(str(cell_value).strip())
        else:
            break
    
    existing_set = set(existing_headers)
    missing_headers = [h for h in expected_headers if h not in existing_set]
    
    if missing_headers:
        next_col = len(existing_headers) + 1
        for header in missing_headers:
            cell = ws.cell(1, next_col, header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            next_col += 1
        log(f"  已补齐缺失字段: {', '.join(missing_headers)}")


def get_existing_images(ws):
    """获取已录入的图片文件名集合"""
    existing = set()
    for row in range(2, ws.max_row + 1):
        link_cell = ws.cell(row, 3).value
        if link_cell and isinstance(link_cell, str):
            if link_cell.startswith('=HYPERLINK'):
                parts = link_cell.split('"')
                if len(parts) >= 2:
                    filename = Path(parts[1]).name
                    existing.add(filename)
    return existing


def list_product_images():
    """列出所有待识别产品图片（主目录下所有图片文件）"""
    images = []
    for item in BASE_DIR.iterdir():
        if item.is_file() and item.suffix.lower() in IMAGE_EXTS:
            images.append(item)
    return sorted(images, key=lambda x: x.name)


def encode_image(image_path):
    """将图片编码为 base64"""
    with open(image_path, "rb") as f:
        return base64.b64encode(f.read()).decode("utf-8")


def identify_product_with_ai(image_path):
    """使用 OpenAI Vision API 识别产品"""
    if not API_KEY:
        log("  警告: 未设置 OPENAI_API_KEY，使用占位符")
        return {
            "prompt": "未配置 API",
            "cn_desc": "未配置 API",
            "en_desc": "API not configured"
        }

    try:
        client = OpenAI(api_key=API_KEY)

        image_data = encode_image(image_path)
        ext = image_path.suffix.lower()
        if ext in [".jpg", ".jpeg"]:
            media_type = "image/jpeg"
        elif ext == ".png":
            media_type = "image/png"
        elif ext == ".webp":
            media_type = "image/webp"
        else:
            media_type = "image/jpeg"

        prompt = """请分析这张产品图片，提供以下信息：

1. **图片提示词**（用于 AI 绘图的详细描述，英文，40-60词）
2. **产品中文描述**（简洁的产品名称和特点，10-15字）
3. **产品英文描述**（Product name and features, 5-10 words）

请严格按照以下 JSON 格式返回：
{
  "prompt": "详细的英文图片提示词",
  "cn_desc": "中文产品描述",
  "en_desc": "English product description"
}"""

        response = client.chat.completions.create(
            model=MODEL,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:{media_type};base64,{image_data}"
                            }
                        },
                        {
                            "type": "text",
                            "text": prompt
                        }
                    ]
                }
            ],
            max_tokens=1024
        )

        response_text = response.choices[0].message.content

        import re
        json_match = re.search(r'\{[^}]+\}', response_text, re.DOTALL)
        if json_match:
            result = json.loads(json_match.group())
            return result
        else:
            log("  警告: 无法解析 API 响应为 JSON")
            return {
                "prompt": response_text[:100],
                "cn_desc": "解析失败",
                "en_desc": "Parse failed"
            }

    except Exception as e:
        log(f"  识别出错: {e}")
        return {
            "prompt": f"识别失败: {str(e)[:50]}",
            "cn_desc": "识别失败",
            "en_desc": "Recognition failed"
        }


def move_image_to_uploaded(image_path):
    """将图片转移到【已上传图片】文件夹"""
    UPLOADED_DIR.mkdir(parents=True, exist_ok=True)
    dest_path = UPLOADED_DIR / image_path.name

    counter = 1
    while dest_path.exists():
        stem = image_path.stem
        ext = image_path.suffix
        dest_path = UPLOADED_DIR / f"{stem}_{counter}{ext}"
        counter += 1

    try:
        shutil.move(str(image_path), str(dest_path))
        log(f"  已转移: {image_path.name} -> 已上传图片/{dest_path.name}")
        return dest_path
    except Exception as e:
        log(f"  转移失败: {image_path.name} - {e}")
        return None


def add_product_to_excel(ws, seq_num, dest_path, result):
    """添加产品到 Excel"""
    row = ws.max_row + 1
    now = datetime.now()

    ws.cell(row, 1, seq_num)
    ws.cell(row, 2, now.strftime("%m/%d"))

    rel_path = f"已上传图片/{dest_path.name}"
    hyperlink = f'=HYPERLINK("{rel_path}", "{dest_path.name}")'
    ws.cell(row, 3, hyperlink)
    ws.cell(row, 3).font = Font(color="0563C1", underline="single")

    ws.cell(row, 4, result["prompt"])
    ws.cell(row, 5, result["cn_desc"])
    ws.cell(row, 6, result["en_desc"])
    ws.cell(row, 7, "否")
    
    # 003 闭环字段默认值
    ws.cell(row, 8, "")  # 记录ID
    ws.cell(row, 9, "002_已识别")  # 流程阶段
    ws.cell(row, 10, "success")  # 识别状态
    ws.cell(row, 11, "pending")  # 前端生成状态
    ws.cell(row, 12, "pending")  # 上传状态
    ws.cell(row, 13, "")  # 上传时间
    ws.cell(row, 14, "")  # 上传目标
    ws.cell(row, 15, "")  # 上传批次号
    ws.cell(row, 16, "")  # 上传结果
    ws.cell(row, 17, "")  # 失败原因
    ws.cell(row, 18, now.strftime("%Y-%m-%d %H:%M:%S"))  # 最近处理时间

    for col in range(1, 19):
        ws.cell(row, col).alignment = Alignment(vertical="center", wrap_text=True)


def refresh_frontend_products():
    """重新生成前端产品展示数据"""
    if not GENERATE_SCRIPT.exists():
        log(f"未找到前端数据生成脚本: {GENERATE_SCRIPT}")
        return False

    try:
        result = subprocess.run(
            [sys.executable, str(GENERATE_SCRIPT)],
            cwd=str(GENERATE_SCRIPT.parent),
            capture_output=True,
            text=True,
            encoding="utf-8",
            check=True,
        )
        log("前端产品数据已自动刷新")
        if result.stdout.strip():
            log(f"generate_products_data 输出: {result.stdout.strip()}")
        return True
    except subprocess.CalledProcessError as e:
        log(f"刷新前端产品数据失败: {e}")
        if e.stdout:
            log(f"stdout: {e.stdout.strip()}")
        if e.stderr:
            log(f"stderr: {e.stderr.strip()}")
        return False


def process_images():
    """处理所有产品图片"""
    log("=" * 60)
    log("开始处理产品图片（OpenAI GPT-5.4 Vision 识别 + 自动转移模式）")

    wb = init_excel()
    ws = wb.active

    existing_images = get_existing_images(ws)
    log(f"已录入图片数量: {len(existing_images)}")

    images = list_product_images()
    log(f"找到待识别图片: {len(images)} 张")

    new_count = 0
    current_seq = ws.max_row

    for image_path in images:
        if image_path.name in existing_images:
            log(f"跳过已录入: {image_path.name}")
            continue

        log(f"处理新图片: {image_path.name}")
        result = identify_product_with_ai(image_path)

        dest_path = move_image_to_uploaded(image_path)
        if dest_path is None:
            log("  跳过该图片（转移失败）")
            continue

        current_seq += 1
        add_product_to_excel(ws, current_seq, dest_path, result)
        new_count += 1

        log(f"  已添加 (序号: {current_seq})")
        log(f"  中文: {result['cn_desc']}")
        log(f"  英文: {result['en_desc']}")

    if new_count > 0:
        wb.save(EXCEL_FILE)
        log(f"已保存 Excel: {EXCEL_FILE}")
        log(f"本次新增: {new_count} 条记录")
        refresh_frontend_products()
    else:
        log("没有新图片需要处理")

    log("=" * 60)
    return new_count


def main():
    if not API_KEY:
        log("警告: 未设置 OPENAI_API_KEY 环境变量")
        log("将创建表格但不进行 AI 识别")

    try:
        process_images()
    except Exception as e:
        log(f"处理出错: {e}")
        import traceback
        log(traceback.format_exc())


if __name__ == "__main__":
    main()
