#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 表头与状态字段管理工具
确保产品识别 Excel 包含完整的 001/002/003 闭环字段
"""

from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# 完整表头定义（按顺序）
HEADER_SCHEMA = [
    # 原有字段
    "序号",
    "添加时间",
    "图片超链接",
    "提示词",
    "中文描述",
    "英文描述",
    "是否上传",
    # 003 闭环字段
    "记录ID",
    "流程阶段",
    "识别状态",
    "前端生成状态",
    "上传状态",
    "上传时间",
    "上传目标",
    "上传批次号",
    "上传结果",
    "失败原因",
    "最近处理时间",
]

# 列宽定义
COLUMN_WIDTHS = {
    "序号": 8,
    "添加时间": 12,
    "图片超链接": 35,
    "提示词": 50,
    "中文描述": 25,
    "英文描述": 40,
    "是否上传": 10,
    "记录ID": 20,
    "流程阶段": 18,
    "识别状态": 12,
    "前端生成状态": 15,
    "上传状态": 12,
    "上传时间": 18,
    "上传目标": 15,
    "上传批次号": 20,
    "上传结果": 25,
    "失败原因": 30,
    "最近处理时间": 18,
}


def ensure_headers(ws):
    """
    确保工作表包含完整表头
    如果缺失字段，自动追加到末尾
    返回：header_map (字段名 -> 列索引)
    """
    # 读取现有表头
    existing_headers = []
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(1, col).value
        if cell_value:
            existing_headers.append(str(cell_value).strip())
        else:
            break

    # 找出缺失字段
    existing_set = set(existing_headers)
    missing_headers = [h for h in HEADER_SCHEMA if h not in existing_set]

    # 追加缺失字段
    if missing_headers:
        next_col = len(existing_headers) + 1
        for header in missing_headers:
            cell = ws.cell(1, next_col, header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            existing_headers.append(header)
            next_col += 1

    # 设置列宽
    for col_idx, header in enumerate(existing_headers, 1):
        if header in COLUMN_WIDTHS:
            col_letter = ws.cell(1, col_idx).column_letter
            ws.column_dimensions[col_letter].width = COLUMN_WIDTHS[header]

    # 构建 header_map
    header_map = {header: idx - 1 for idx, header in enumerate(existing_headers, 1)}
    return header_map


def get_default_values_for_new_row():
    """
    返回新增记录时的默认字段值
    """
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return {
        "记录ID": "",
        "流程阶段": "002_已识别",
        "识别状态": "success",
        "前端生成状态": "pending",
        "上传状态": "pending",
        "上传时间": "",
        "上传目标": "",
        "上传批次号": "",
        "上传结果": "",
        "失败原因": "",
        "最近处理时间": now,
    }


def apply_defaults_to_row(ws, row_number, header_map, defaults):
    """
    对指定行应用默认值（仅当单元格为空时）
    """
    for field, default_value in defaults.items():
        if field in header_map:
            col_idx = header_map[field] + 1
            cell = ws.cell(row_number, col_idx)
            if not cell.value:
                cell.value = default_value
                cell.alignment = Alignment(vertical="center", wrap_text=True)


def backfill_existing_rows(ws, header_map):
    """
    对已有记录补齐缺失的状态字段默认值
    """
    defaults = get_default_values_for_new_row()
    for row in range(2, ws.max_row + 1):
        apply_defaults_to_row(ws, row, header_map, defaults)


def ensure_excel_schema(excel_path):
    """
    确保 Excel 文件包含完整的 003 闭环字段
    如果文件不存在，不做任何操作
    """
    excel_path = Path(excel_path)
    if not excel_path.exists():
        return False

    wb = load_workbook(excel_path)
    ws = wb[wb.sheetnames[0]]

    # 确保表头完整
    header_map = ensure_headers(ws)

    # 补齐已有记录的默认值
    backfill_existing_rows(ws, header_map)

    wb.save(excel_path)
    return True


if __name__ == "__main__":
    # 测试用例
    excel_file = Path(r"C:\Users\Win11\Desktop\XCSOURCE\assets\products\产品识别.xlsx")
    if ensure_excel_schema(excel_file):
        print(f"✓ 已更新 Excel 表头与字段: {excel_file}")
    else:
        print(f"✗ Excel 文件不存在: {excel_file}")
