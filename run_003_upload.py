#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
003 产品上传执行与状态回写脚本

职责：
- 只读取 001/002 已产出的 Excel 记录
- 校验 003 进入条件
- 执行上传动作（当前支持 simulate / manual-mark / local-copy）
- 回写 Excel 与 upload_state.json
- 记录批次日志
- 生成批次 manifest
- 可选：提交并推送 003 产物到 GitHub

注意：
- 不修改 `是否上传` 的语义
- 不重复做 001/002 的采集、识别、前端生成工作
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
import subprocess
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Alignment

PROJECT_ROOT = Path(r"C:\Users\Win11\Desktop\XCSOURCE")
PRODUCTS_DIR = PROJECT_ROOT / "assets" / "products"
EXCEL_PATH = PRODUCTS_DIR / "产品识别.xlsx"
UPLOAD_STATE_FILE = PRODUCTS_DIR / "upload_state.json"
LOG_FILE = PRODUCTS_DIR / "003-upload.log"
UPLOAD_OUTBOX_DIR = PRODUCTS_DIR / "uploaded_by_003"
DEFAULT_UPLOAD_TARGET = "github_pages"
DEFAULT_GIT_REMOTE = "origin"
DEFAULT_GIT_BRANCH = "master"
HYPERLINK_RE = re.compile(r'=HYPERLINK\("([^"]+)"\s*,\s*"([^"]*)"\)', re.IGNORECASE)


@dataclass
class UploadResult:
    status: str
    result: str = ""
    error: str = ""
    remote_path: str = ""


@dataclass
class ProductRecord:
    row: int
    sequence: str
    record_id: str
    image_link_raw: str
    image_path: Path
    image_name: str
    english_description: str
    upload_status: str
    recognition_status: str


def now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def now_compact() -> str:
    return datetime.now().strftime("%Y%m%d-%H%M%S")


def log(msg: str) -> None:
    ts = now_str()
    line = f"[{ts}] {msg}"
    print(line)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(line + "\n")


def load_upload_state() -> Dict[str, Dict]:
    if not UPLOAD_STATE_FILE.exists():
        return {}
    try:
        return json.loads(UPLOAD_STATE_FILE.read_text(encoding="utf-8"))
    except Exception as e:
        log(f"加载上传状态文件失败，改用空状态: {e}")
        return {}


def save_upload_state(state: Dict[str, Dict]) -> None:
    UPLOAD_STATE_FILE.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")


def get_header_map(ws) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(1, col).value
        if value:
            headers[str(value).strip()] = col
    return headers


def require_headers(header_map: Dict[str, int], fields: List[str]) -> None:
    missing = [field for field in fields if field not in header_map]
    if missing:
        raise KeyError(f"Excel 缺少必要字段: {', '.join(missing)}")


def parse_hyperlink(raw_value: object) -> Tuple[Optional[str], Optional[str]]:
    if raw_value is None:
        return None, None
    text = str(raw_value).strip()
    match = HYPERLINK_RE.fullmatch(text)
    if match:
        return match.group(1).strip(), match.group(2).strip()
    return None, None


def resolve_image_path(link_target: str) -> Path:
    link_path = Path(link_target.replace("\\", "/"))
    if link_path.is_absolute():
        return link_path
    return PRODUCTS_DIR / link_path


def normalize_record_id(value: object, sequence: object, row_num: int) -> str:
    text = str(value).strip() if value is not None else ""
    if text:
        return text
    seq = str(sequence).strip() if sequence is not None else str(row_num - 1)
    seq = seq or str(row_num - 1)
    return f"PRD-{datetime.now().strftime('%Y%m%d')}-{int(float(seq)):04d}" if str(seq).replace('.', '', 1).isdigit() else f"PRD-{datetime.now().strftime('%Y%m%d')}-ROW{row_num}"


def read_pending_records(ws, header_map: Dict[str, int]) -> List[ProductRecord]:
    require_headers(
        header_map,
        ["序号", "图片超链接", "英文描述", "识别状态", "上传状态", "记录ID"],
    )

    records: List[ProductRecord] = []
    for row_num in range(2, ws.max_row + 1):
        try:
            sequence = ws.cell(row_num, header_map["序号"]).value
            image_link_raw = ws.cell(row_num, header_map["图片超链接"]).value
            english_description = ws.cell(row_num, header_map["英文描述"]).value
            recognition_status = ws.cell(row_num, header_map["识别状态"]).value
            upload_status = ws.cell(row_num, header_map["上传状态"]).value
            record_id_raw = ws.cell(row_num, header_map["记录ID"]).value

            description = str(english_description).strip() if english_description is not None else ""
            recog = str(recognition_status).strip() if recognition_status is not None else ""
            upload = str(upload_status).strip() if upload_status is not None else ""

            if not description:
                continue
            if not image_link_raw:
                continue
            if recog != "success":
                continue
            if upload not in ["pending", "retry"]:
                continue

            link_target, display_text = parse_hyperlink(image_link_raw)
            if not link_target:
                log(f"跳过第 {row_num} 行：图片超链接格式无效")
                continue

            image_path = resolve_image_path(link_target)
            if not image_path.exists() or not image_path.is_file():
                log(f"跳过第 {row_num} 行：图片不存在 -> {image_path}")
                continue

            record_id = normalize_record_id(record_id_raw, sequence, row_num)
            records.append(
                ProductRecord(
                    row=row_num,
                    sequence=str(sequence).strip() if sequence is not None else "",
                    record_id=record_id,
                    image_link_raw=str(image_link_raw),
                    image_path=image_path,
                    image_name=display_text or image_path.name,
                    english_description=description,
                    upload_status=upload,
                    recognition_status=recog,
                )
            )
        except Exception as e:
            log(f"读取第 {row_num} 行失败: {e}")
    return records


def set_cell_if_exists(ws, header_map: Dict[str, int], row: int, field: str, value: object) -> None:
    col = header_map.get(field)
    if col:
        ws.cell(row, col, value)


def apply_alignment(ws, header_map: Dict[str, int], row: int) -> None:
    for col in header_map.values():
        ws.cell(row, col).alignment = Alignment(vertical="center", wrap_text=True)


def mark_record_uploading(ws, header_map: Dict[str, int], record: ProductRecord, batch_id: str, target: str) -> None:
    ts = now_str()
    set_cell_if_exists(ws, header_map, record.row, "记录ID", record.record_id)
    set_cell_if_exists(ws, header_map, record.row, "流程阶段", "003_上传中")
    set_cell_if_exists(ws, header_map, record.row, "上传状态", "uploading")
    set_cell_if_exists(ws, header_map, record.row, "上传目标", target)
    set_cell_if_exists(ws, header_map, record.row, "上传批次号", batch_id)
    set_cell_if_exists(ws, header_map, record.row, "最近处理时间", ts)
    apply_alignment(ws, header_map, record.row)


def execute_upload_local_copy(record: ProductRecord, target: str, batch_id: str) -> UploadResult:
    try:
        dest_dir = UPLOAD_OUTBOX_DIR / batch_id / record.record_id
        dest_dir.mkdir(parents=True, exist_ok=True)

        dest_image = dest_dir / record.image_path.name
        shutil.copy2(record.image_path, dest_image)

        metadata = {
            "record_id": record.record_id,
            "sequence": record.sequence,
            "target": target,
            "source_image": str(record.image_path),
            "uploaded_image": str(dest_image),
            "english_description": record.english_description,
            "uploaded_at": now_str(),
            "mode": "local-copy",
        }
        metadata_path = dest_dir / "upload_metadata.json"
        metadata_path.write_text(json.dumps(metadata, ensure_ascii=False, indent=2), encoding="utf-8")

        return UploadResult(
            status="success",
            result=f"已复制到 003 上传目录: {dest_image}",
            error="",
            remote_path=str(dest_dir),
        )
    except Exception as e:
        return UploadResult(status="failed", result="", error=f"local-copy 失败: {e}")


def execute_upload(record: ProductRecord, mode: str, target: str, batch_id: str) -> UploadResult:
    try:
        if mode == "simulate":
            log(f"  [simulate] {record.record_id} -> {record.image_name}")
            return UploadResult(
                status="success",
                result=f"模拟上传成功: {record.image_name} -> {target}",
                error="",
            )

        if mode == "manual-mark":
            log(f"  [manual-mark] {record.record_id} -> {record.image_name}")
            return UploadResult(
                status="success",
                result=f"人工确认已上传: {record.image_name} -> {target}",
                error="",
            )

        if mode == "local-copy":
            log(f"  [local-copy] {record.record_id} -> {record.image_name}")
            return execute_upload_local_copy(record, target=target, batch_id=batch_id)

        return UploadResult(status="failed", result="", error=f"不支持的上传模式: {mode}")
    except Exception as e:
        return UploadResult(status="failed", result="", error=str(e))


def writeback_excel(ws, header_map: Dict[str, int], record: ProductRecord, result: UploadResult, batch_id: str, target: str) -> None:
    ts = now_str()
    set_cell_if_exists(ws, header_map, record.row, "记录ID", record.record_id)
    set_cell_if_exists(ws, header_map, record.row, "上传状态", result.status)
    set_cell_if_exists(ws, header_map, record.row, "上传目标", target)
    set_cell_if_exists(ws, header_map, record.row, "上传批次号", batch_id)
    set_cell_if_exists(ws, header_map, record.row, "上传结果", result.result)
    set_cell_if_exists(ws, header_map, record.row, "最近处理时间", ts)

    if result.status == "success":
        set_cell_if_exists(ws, header_map, record.row, "流程阶段", "003_已回写")
        set_cell_if_exists(ws, header_map, record.row, "上传时间", ts)
        set_cell_if_exists(ws, header_map, record.row, "失败原因", "")
    elif result.status == "failed":
        set_cell_if_exists(ws, header_map, record.row, "流程阶段", "003_上传失败")
        set_cell_if_exists(ws, header_map, record.row, "失败原因", result.error)
    elif result.status == "retry":
        set_cell_if_exists(ws, header_map, record.row, "流程阶段", "003_待上传")
        set_cell_if_exists(ws, header_map, record.row, "失败原因", result.error)
    else:
        set_cell_if_exists(ws, header_map, record.row, "流程阶段", "003_已回写")

    apply_alignment(ws, header_map, record.row)


def writeback_json(state: Dict[str, Dict], record: ProductRecord, result: UploadResult, batch_id: str, target: str) -> None:
    ts = now_str()
    state[record.record_id] = {
        "sequence": record.sequence,
        "image_name": record.image_name,
        "image_path": str(record.image_path),
        "english_description": record.english_description,
        "upload_status": result.status,
        "stage": "003_已回写" if result.status == "success" else ("003_待上传" if result.status == "retry" else "003_上传失败"),
        "upload_target": target,
        "uploaded_at": ts if result.status == "success" else "",
        "batch_id": batch_id,
        "result": result.result,
        "error": result.error,
        "remote_path": result.remote_path,
        "last_processed_at": ts,
    }


def build_manifest_path(batch_id: str) -> Path:
    return UPLOAD_OUTBOX_DIR / batch_id / "manifest.json"


def write_batch_manifest(
    batch_id: str,
    mode: str,
    target: str,
    records: List[ProductRecord],
    processed_items: List[Dict],
    git_summary: Dict[str, object],
) -> Path:
    batch_dir = UPLOAD_OUTBOX_DIR / batch_id
    batch_dir.mkdir(parents=True, exist_ok=True)
    manifest_path = build_manifest_path(batch_id)

    success_count = sum(1 for item in processed_items if item["status"] == "success")
    failed_count = sum(1 for item in processed_items if item["status"] == "failed")
    other_count = len(processed_items) - success_count - failed_count

    manifest = {
        "batch_id": batch_id,
        "generated_at": now_str(),
        "mode": mode,
        "target": target,
        "total_records": len(records),
        "summary": {
            "success": success_count,
            "failed": failed_count,
            "other": other_count,
        },
        "git": git_summary,
        "items": processed_items,
    }
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    return manifest_path


def run_git_command(args: List[str]) -> subprocess.CompletedProcess:
    return subprocess.run(args, cwd=PROJECT_ROOT, capture_output=True, text=True, encoding="utf-8", errors="replace")


def get_tracked_paths(batch_id: str) -> List[str]:
    tracked_paths = [
        str(EXCEL_PATH.relative_to(PROJECT_ROOT)),
        str(UPLOAD_STATE_FILE.relative_to(PROJECT_ROOT)),
        str(LOG_FILE.relative_to(PROJECT_ROOT)),
    ]
    uploaded_batch_dir = UPLOAD_OUTBOX_DIR / batch_id
    if uploaded_batch_dir.exists():
        tracked_paths.append(str(uploaded_batch_dir.relative_to(PROJECT_ROOT)))
    return tracked_paths


def git_precheck(remote: str, branch: str) -> Tuple[bool, Dict[str, object]]:
    summary: Dict[str, object] = {
        "ok": True,
        "current_branch": "",
        "remote_exists": False,
        "branch_matches": False,
        "dirty_paths_count": 0,
        "dirty_paths_preview": [],
        "warnings": [],
        "errors": [],
    }

    branch_result = run_git_command(["git", "branch", "--show-current"])
    if branch_result.returncode != 0:
        summary["ok"] = False
        summary["errors"].append(f"无法获取当前分支: {branch_result.stderr.strip() or branch_result.stdout.strip()}")
        return False, summary

    current_branch = (branch_result.stdout or "").strip()
    summary["current_branch"] = current_branch
    summary["branch_matches"] = current_branch == branch
    if current_branch != branch:
        summary["warnings"].append(f"当前分支为 {current_branch}，目标推送分支为 {branch}")

    remote_result = run_git_command(["git", "remote"])
    if remote_result.returncode != 0:
        summary["ok"] = False
        summary["errors"].append(f"无法读取 git remote: {remote_result.stderr.strip() or remote_result.stdout.strip()}")
        return False, summary

    remotes = [line.strip() for line in (remote_result.stdout or "").splitlines() if line.strip()]
    summary["remote_exists"] = remote in remotes
    if remote not in remotes:
        summary["ok"] = False
        summary["errors"].append(f"git remote 不存在: {remote}")

    status_result = run_git_command(["git", "status", "--short"])
    if status_result.returncode == 0:
        dirty_paths = [line.rstrip() for line in (status_result.stdout or "").splitlines() if line.strip()]
        summary["dirty_paths_count"] = len(dirty_paths)
        summary["dirty_paths_preview"] = dirty_paths[:20]
        if dirty_paths:
            summary["warnings"].append(f"工作区存在 {len(dirty_paths)} 条未清理改动；脚本仍只会提交 003 相关文件")
    else:
        summary["warnings"].append(f"无法读取 git status: {status_result.stderr.strip() or status_result.stdout.strip()}")

    return bool(summary["ok"]), summary


def git_commit_and_push(batch_id: str, mode: str, target: str, remote: str, branch: str, commit_only: bool) -> Tuple[bool, str, Dict[str, object]]:
    tracked_paths = get_tracked_paths(batch_id)
    git_details: Dict[str, object] = {
        "action": "commit-only" if commit_only else "commit-and-push",
        "tracked_paths": tracked_paths,
        "commit_message": "",
        "commit_sha": "",
        "push_skipped": commit_only,
    }

    precheck_ok, precheck = git_precheck(remote=remote, branch=branch)
    git_details["precheck"] = precheck
    if not precheck_ok:
        return False, "git 预检查失败", git_details

    add_cmd = ["git", "add", "--"] + tracked_paths
    add_result = run_git_command(add_cmd)
    git_details["add"] = {
        "returncode": add_result.returncode,
        "stdout": (add_result.stdout or "").strip(),
        "stderr": (add_result.stderr or "").strip(),
    }
    if add_result.returncode != 0:
        return False, f"git add 失败: {add_result.stderr.strip() or add_result.stdout.strip()}", git_details

    diff_result = run_git_command(["git", "diff", "--cached", "--quiet", "--"] + tracked_paths)
    git_details["diff_cached"] = {"returncode": diff_result.returncode}
    if diff_result.returncode == 0:
        return True, "没有可提交的 003 变更", git_details
    if diff_result.returncode not in (0, 1):
        return False, f"git diff --cached 检查失败: {diff_result.stderr.strip() or diff_result.stdout.strip()}", git_details

    commit_message = f"feat: 003 上传批次 {batch_id}\n\n- mode: {mode}\n- target: {target}\n- state: Excel + upload_state.json + upload log + manifest"
    git_details["commit_message"] = commit_message

    commit_result = run_git_command(["git", "commit", "-m", commit_message])
    git_details["commit"] = {
        "returncode": commit_result.returncode,
        "stdout": (commit_result.stdout or "").strip(),
        "stderr": (commit_result.stderr or "").strip(),
    }
    if commit_result.returncode != 0:
        return False, f"git commit 失败: {commit_result.stderr.strip() or commit_result.stdout.strip()}", git_details

    rev_result = run_git_command(["git", "rev-parse", "HEAD"])
    if rev_result.returncode == 0:
        git_details["commit_sha"] = (rev_result.stdout or "").strip()

    if commit_only:
        return True, f"已完成本地提交: {git_details['commit_sha'] or 'latest'}", git_details

    push_result = run_git_command(["git", "push", remote, branch])
    git_details["push"] = {
        "returncode": push_result.returncode,
        "stdout": (push_result.stdout or "").strip(),
        "stderr": (push_result.stderr or "").strip(),
    }
    if push_result.returncode != 0:
        return False, f"git push 失败: {push_result.stderr.strip() or push_result.stdout.strip()}", git_details

    return True, f"已推送到 {remote}/{branch}", git_details


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="003 产品上传执行与状态回写")
    parser.add_argument("--mode", choices=["simulate", "manual-mark", "local-copy"], default="simulate", help="上传执行模式")
    parser.add_argument("--target", default=DEFAULT_UPLOAD_TARGET, help="上传目标标识")
    parser.add_argument("--limit", type=int, default=0, help="仅处理前 N 条待上传记录，0 表示全部")
    parser.add_argument("--git-push", action="store_true", help="处理成功后自动提交并推送 003 产物到 GitHub")
    parser.add_argument("--git-commit-only", action="store_true", help="只执行 git add + commit，不执行 git push")
    parser.add_argument("--git-remote", default=DEFAULT_GIT_REMOTE, help="git push 的 remote 名称")
    parser.add_argument("--git-branch", default=DEFAULT_GIT_BRANCH, help="git push 的分支名称")
    return parser.parse_args()


def run_003_upload(mode: str, target: str, limit: int, git_push: bool, git_commit_only: bool, git_remote: str, git_branch: str) -> int:
    log("=" * 60)
    log(f"003 产品上传执行开始 | mode={mode} | target={target}")

    if git_push and git_commit_only:
        log("错误: --git-push 与 --git-commit-only 不能同时使用")
        return 1

    if not EXCEL_PATH.exists():
        log(f"错误: Excel 文件不存在 - {EXCEL_PATH}")
        return 1

    batch_id = f"UP-{now_compact()}"
    log(f"批次号: {batch_id}")

    wb = load_workbook(EXCEL_PATH)
    ws = wb[wb.sheetnames[0]]
    header_map = get_header_map(ws)

    records = read_pending_records(ws, header_map)
    if limit > 0:
        records = records[:limit]
    log(f"找到待上传记录: {len(records)} 条")

    if not records:
        log("没有待上传记录")
        log("=" * 60)
        return 0

    upload_state = load_upload_state()

    success_count = 0
    failed_count = 0
    skipped_count = 0
    processed_items: List[Dict] = []
    git_enabled = git_push or git_commit_only
    git_summary: Dict[str, object] = {
        "enabled": git_enabled,
        "action": "commit-only" if git_commit_only else ("commit-and-push" if git_push else "none"),
        "status": "not-run",
        "message": "",
        "remote": git_remote if git_enabled else "",
        "branch": git_branch if git_enabled else "",
        "details": {},
    }

    for record in records:
        log(f"处理记录: {record.record_id} | 序号 {record.sequence} | {record.english_description}")

        mark_record_uploading(ws, header_map, record, batch_id, target)
        wb.save(EXCEL_PATH)

        result = execute_upload(record, mode=mode, target=target, batch_id=batch_id)

        writeback_excel(ws, header_map, record, result, batch_id, target)
        writeback_json(upload_state, record, result, batch_id, target)
        wb.save(EXCEL_PATH)
        save_upload_state(upload_state)

        processed_items.append({
            "record_id": record.record_id,
            "sequence": record.sequence,
            "row": record.row,
            "image_name": record.image_name,
            "image_path": str(record.image_path),
            "english_description": record.english_description,
            "status": result.status,
            "result": result.result,
            "error": result.error,
            "remote_path": result.remote_path,
        })

        if result.status == "success":
            success_count += 1
            log("  ✓ 上传成功")
        elif result.status == "failed":
            failed_count += 1
            log(f"  ✗ 上传失败: {result.error}")
        else:
            skipped_count += 1
            log(f"  - 状态: {result.status}")

    manifest_path = write_batch_manifest(
        batch_id=batch_id,
        mode=mode,
        target=target,
        records=records,
        processed_items=processed_items,
        git_summary=git_summary,
    )

    if git_enabled and failed_count == 0:
        ok, message, details = git_commit_and_push(
            batch_id=batch_id,
            mode=mode,
            target=target,
            remote=git_remote,
            branch=git_branch,
            commit_only=git_commit_only,
        )
        git_summary["status"] = "success" if ok else "failed"
        git_summary["message"] = message
        git_summary["details"] = details
        if ok:
            if git_commit_only:
                log(f"Git 本地提交完成: {message}")
            else:
                log(f"Git 推送完成: {message}")
        else:
            failed_count += 1
            log(f"Git 执行失败: {message}")
        manifest_path = write_batch_manifest(
            batch_id=batch_id,
            mode=mode,
            target=target,
            records=records,
            processed_items=processed_items,
            git_summary=git_summary,
        )
    elif git_enabled:
        git_summary["status"] = "skipped"
        git_summary["message"] = "存在失败记录，跳过 git 执行"
        manifest_path = write_batch_manifest(
            batch_id=batch_id,
            mode=mode,
            target=target,
            records=records,
            processed_items=processed_items,
            git_summary=git_summary,
        )

    log("")
    log(f"批次 {batch_id} 完成:")
    log(f"  成功: {success_count}")
    log(f"  失败: {failed_count}")
    log(f"  跳过/其他: {skipped_count}")
    log(f"  总计: {len(records)}")
    log(f"  Excel: {EXCEL_PATH}")
    log(f"  JSON: {UPLOAD_STATE_FILE}")
    log(f"  Manifest: {manifest_path}")
    if mode == "local-copy":
        log(f"  上传目录: {UPLOAD_OUTBOX_DIR}")
    if git_enabled:
        log(f"  Git: action={git_summary['action']}, remote={git_remote}, branch={git_branch}, status={git_summary['status']}")
    log("=" * 60)
    return 0 if failed_count == 0 else 2


def main() -> None:
    args = parse_args()
    try:
        raise SystemExit(
            run_003_upload(
                mode=args.mode,
                target=args.target,
                limit=args.limit,
                git_push=args.git_push,
                git_commit_only=args.git_commit_only,
                git_remote=args.git_remote,
                git_branch=args.git_branch,
            )
        )
    except SystemExit:
        raise
    except Exception as e:
        log(f"003 执行出错: {e}")
        import traceback
        log(traceback.format_exc())
        raise SystemExit(1)


if __name__ == "__main__":
    main()
